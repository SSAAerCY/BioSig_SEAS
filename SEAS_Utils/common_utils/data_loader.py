#!/usr/bin/env python
#
# Copyright (C) 2017 - Massachusetts Institute of Technology (MIT)
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

"""
Misc. Data Loading functions


Should I put the hitran data reader here as well?
i mean, after all the user should want control what data line list data
it's reading in from...

"""

import os
import sys
import json
import numpy as np
from openpyxl import load_workbook, Workbook

DIR = os.path.abspath(os.path.dirname(__file__))
sys.path.insert(0, os.path.join(DIR, '../..'))

import SEAS_Utils.common_utils.db_management2 as dbm
import SEAS_Utils.common_utils.jdx_Reader as jdx

from SEAS_Utils.common_utils.data_saver import check_path_exist, check_file_exist
from SEAS_Main.atmosphere_effects.biosig_molecule import biosig_interpolate


def two_column_file_loader(path,spliter=None,type="float",skip=0):
    """
    load data from files that contains only two column
    """
    
    with open(path) as data_file:   
        data = [x.split(spliter) for x in data_file.read().split("\n")[skip:]]
        if data[-1] == []:
            data = data[:-1]        
        xdata,ydata = [list(x) for x  in zip(*data)]     
        if type == "float":
            return np.array(xdata,dtype=np.float),np.array(ydata,dtype=np.float)
        elif type == "int":
            return np.array(xdata,dtype=np.int),np.array(ydata,dtype=np.int)
        elif type == "mixed":
            return xdata,ydata
        else:
            return xdata,ydata

def two_column_chunk_file_loader(path,spliter=None,chunk_splitter="\n",type="float",skip=0):
    pass
      
def multi_column_file_loader(path,spliter=None,type="float",skip=0):
    """
    load data from files that contains multiple columns
    """
    
    with open(path) as data_file:   
        data = [x.split(spliter) for x in data_file.read().split("\n")[skip:]]
        if data[-1] == []:
            data = data[:-1]
        
        data = [list(x) for x  in zip(*data)]    
        
        if type == "float":
            return np.array(data,dtype=np.float)
        elif type == "int":
            return np.array(data,dtype=np.int)
        elif type == "mixed":
            return data
    
def json_loader(path):
    
    with open(path) as data_file:
        data = json.load(data_file)

    return data

def load_npy(savepath, savename):
    
    if ".npy" in savename:
        save = os.path.join(savepath, savename)
    else:
        save = os.path.join(savepath, savename)+".npy"
        

    if check_file_exist(save):
        return np.load(save)
    else:
        print "File %s Does not exist"%save
        return []


def molecule_cross_section_loader(inputs, cross_db, molecule):
    """
    Need to move this to the SEAS_Main
    """
    
    Pgrid = inputs["Simulation_Control"]["P_Grid"]
    Tgrid = inputs["Simulation_Control"]["T_Grid"]
    numin = inputs["Spectra"]["Numin"]
    numax = inputs["Spectra"]["Numax"]


    data = [[0 for y in range(len(Pgrid))] for x in range(len(Tgrid))]
    for i,T in enumerate(Tgrid):
        for j,P in enumerate(Pgrid):
            
            print molecule,P,T,numin,numax
            result = cross_db.c.execute("SELECT nu, coef FROM {} WHERE P={} AND T={} AND nu>={} AND nu<={} ORDER BY nu".format(molecule,P,T,numin,numax))
            fetch = np.array(result.fetchall()).T
            data[i][j] = fetch[1]
            
    nu = fetch[0]

    
    return nu, data

def molecule_cross_section_loader2(inputs, db_dir, molecule):
    """
    Need to move this to the SEAS_Main
    """

    T_grid = inputs["Simulation_Control"]["T_Grid"]
    P_grid = inputs["Simulation_Control"]["P_Grid"]

    kwargs = {"dir"        :db_dir,
              "db_name"    :"cross_section_Simulation_%s.db"%molecule,
              "user"       :"azariven",
              "DEBUG"      :False,
              "REMOVE"     :True,
              "BACKUP"     :False,
              "OVERWRITE"  :True}
    
    cross_db = dbm.database(**kwargs)  
    cross_db.access_db()  

    data = [[0 for y in range(len(T_grid))] for x in range(len(P_grid))]
    for j,P in enumerate(P_grid):
        for i,T in enumerate(T_grid):
            table_name = "T%sP%s"%(i,j)
            result = cross_db.c.execute("SELECT * FROM {} ORDER BY nu".format(table_name))
            fetch = np.array(result.fetchall()).T
            data[j][i] = fetch[1]
            
    nu = fetch[0]

    
    return nu, data

def exomol_cross_section_loader(inputs, nu, Exomol_DB_DIR, molecule):
    
    

    T_grid = inputs["Simulation_Control"]["T_Grid"]
    P_grid = inputs["Simulation_Control"]["P_Grid"]


    kwargs = {"dir"        :Exomol_DB_DIR,
              "db_name"    :"Exomol_%s.db"%molecule,
              "user"       :"azariven",
              "DEBUG"      :False,
              "REMOVE"     :True,
              "BACKUP"     :False,
              "OVERWRITE"  :True}
    
    cross_db = dbm.database(**kwargs)  
    cross_db.access_db()  


    data = [[0 for y in range(len(T_grid))] for x in range(len(P_grid))]
    for j,P in enumerate(P_grid):
        for i,T in enumerate(T_grid):
            table_name = "T%sP%s"%(i,j)
            result = cross_db.c.execute("SELECT * FROM {} ORDER BY nu".format(table_name))
            fetch = np.array(result.fetchall()).T

            x1 = np.array(fetch[0],dtype="float")
            y1 = np.array(fetch[1],dtype="float")

            y2 = biosig_interpolate(x1,y1,nu,"C")
            data[j][i] = y2

    return nu, data


def HITRAN_Line_List_reference():
    
    from SEAS_Utils.common_utils.DIRs import HITRAN_Molecule_List
    
    molecule = open(HITRAN_Molecule_List).read().split("\n")
    
    component = []
    for i in range(len(molecule)):
        component.append([i+1,1,1])
        
    return molecule,component



def NIST_Smile_List():


    kwargs = {"db_name":"Molecule_DB.db",
              "user":"azariven",
              "dir":"../../input/molecule_info",
              "DEBUG":False,"REMOVE":False,"BACKUP":False,"OVERWRITE":False}
    
    cross_db = dbm.database(**kwargs)   
    cross_db.access_db()   

    """
    cmd = 'SELECT ID.smiles, ID.inchikey FROM ID,Spectra \
            WHERE ID.inchikey=Spectra.inchikey AND Spectra.has_spectra="Y"'
    ""
    if info == "Smiles":
        cmd = "SELECT SMILES FROM Spectra WHERE IS_Gas='Y'"
    elif info == "CAS":
        cmd = "SELECT CAS FROM Spectra WHERE IS_Gas='Y'"
    elif info == "Inchikey":
        cmd = "SELECT Inchikey FROM Spectra WHERE IS_Gas='Y'"
    else:
        print "Unknown Info Type, Simulation Terminated"
        sys.exit()
    """
    cmd = 'SELECT Smiles, Inchikey, CAS FROM Spectra WHERE Is_Gas="Y"'
    
    
    result = cross_db.c.execute(cmd)
    data = np.array(result.fetchall()).T
    
    
    smiles = data[0]
    inchikeys = data[1]
    #CAS = data[2]

    return smiles, inchikeys




class Excel_Loader():
    
    def __init__(self, path, name, sheetname="sheet"):
        
        #check_path_exist(path)
        #check_file_exist
        
        self.path = path
        self.name = name
        self.sheetname = sheetname
        
        self.file = os.path.join(self.path,self.name)
        
    
    def create(self):
        
        self.WBI = Workbook()
    
    
    def load(self):
        
        self.WBI = load_workbook(self.file)
        input_data = self.WBI[self.sheetname]
        

class database_loader():
    "TBD, keep this here."
    
    def __init__(self):
        
        pass
        


