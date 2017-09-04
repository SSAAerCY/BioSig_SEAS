"""
Convert the all small molecule information into molecule.db for accessing molecular information
so that it can be used for NIST Spectra Scraping

"""


import os
import sys
from openpyxl import Workbook, load_workbook

DIR = os.path.abspath(os.path.dirname(__file__))
sys.path.insert(0, os.path.join(DIR, '../..'))

import SEAS_Utils as utils
from SEAS_Utils.common_utils.DIRs import molecule_info
import SEAS_Utils.common_utils.db_management2 as dbm




def create_Molecule_DB():

    inputfilename = os.path.join(molecule_info,"ASM4.3_Lite.xlsx")
    inputsheetname = "Database"

    WBI = load_workbook(inputfilename)
    data_sheet_ranges = WBI[inputsheetname]
    
    Header = []
    col = ord("A")
    while True:
        cell_value = data_sheet_ranges['%s1'%chr(col)].value
        if cell_value == None:
            break
        Header.append(cell_value.replace(" ","_"))
        col+=1
        

    Data = []
    num = 2
    while True:
        Smile            = utils.to_str(data_sheet_ranges['A%d'%num].value)
        if Smile == None:
            break

        compatible_Smile = utils.to_str(data_sheet_ranges['B%d'%num].value)
        database_number  = utils.to_int(data_sheet_ranges['C%d'%num].value)
        molecular_weight = utils.to_float(data_sheet_ranges['D%d'%num].value)
        formula          = utils.to_str(data_sheet_ranges['E%d'%num].value)
        iupac_name       = utils.to_str(data_sheet_ranges['F%d'%num].value)
        number_atom      = utils.to_int(data_sheet_ranges['G%d'%num].value)
        number_NH_atom   = utils.to_int(data_sheet_ranges['H%d'%num].value)
        
        try:
            inchi        = utils.to_str(data_sheet_ranges['I%d'%num].value).split("=")[-1]
        except:
            inchi        = None
        
        try:
            inchikey     = utils.to_str(data_sheet_ranges['J%d'%num].value).split("=")[-1]
        except:
            inchikey     = None            
        
        try:
            boiling_point    = utils.to_float(data_sheet_ranges['K%d'%num].value)
        except:
            boiling_point    = None
    
        produced_by_life = utils.to_str(data_sheet_ranges['L%d'%num].value)
        cas_registry     = utils.to_int(data_sheet_ranges['M%d'%num].value)
        
        Data.append([Smile,compatible_Smile,database_number,molecular_weight,
                     formula,iupac_name,number_atom,number_NH_atom,
                     inchi,inchikey,boiling_point,produced_by_life])
        num+=1


    kwargs = {  "dir"        :molecule_info,
                "db_name"    :"Molecule_DB2.db",
                "user"       :"Azariven",
                "DEBUG"      :False,
                "REMOVE"     :True,
                "BACKUP"     :False,
                "OVERWRITE"  :True}

    cross_db = dbm.database(**kwargs)
    
    if cross_db.is_db():
        print "Access DB"
        cross_db.access_db()
    else:
        print "Create DB"
        cross_db.create_db()


    table_name = "ID"
    columns = [(str(x),"text") for x in Header]
    print columns

    if cross_db.check_table_exist(table_name) == False:
        cross_db.create_table(table_name, *columns)
    else:
        cross_db.delete_table(table_name)
        cross_db.create_table(table_name, *columns)
    
    for info in Data:
        cross_db._insert_data_single_dev(table_name, info)
    
    cross_db.conn.commit() 
    
    
def create_Spectra_DB():
    
    inputfilename = os.path.join(molecule_info,"ASM4.3_Lite.xlsx")
    inputsheetname = "Has Spectra"    
    
    WBI = load_workbook(inputfilename)
    data_sheet_ranges = WBI[inputsheetname]    

    kwargs = {"dir":molecule_info,
              "db_name":"Molecule_DB2.db",
              "user":"azariven",
              "DEBUG":False,"REMOVE":False,"BACKUP":False,"OVERWRITE":False}
    
    cross_db = dbm.database(**kwargs)   
    cross_db.access_db()   

    table_name = "Spectra"
    Header = ["Smiles", "CAS", "Inchikey","Is_Spectra", "IS_Gas", "Path"]
    columns = [(str(x),"text") for x in Header]
    
    num = 2
    Data = []
    while True:
        
        Smile            = utils.to_str(data_sheet_ranges['A%d'%num].value)
        CAS              = utils.to_str(data_sheet_ranges['M%d'%num].value)
        if Smile == None or CAS == None:
            break          
        
        Inchikey         = utils.to_str(data_sheet_ranges['J%d'%num].value).split("=")[-1]
        Is_Spectra       = utils.to_str(data_sheet_ranges['N%d'%num].value)
        Is_Gas           = utils.to_str(data_sheet_ranges['O%d'%num].value)
        Path             = os.path.join(CAS,"%s_0.jdx"%CAS)
    
        Data.append([Smile,CAS,Inchikey,Is_Spectra,Is_Gas,Path])
        
  
        
        num+=1
        
        
    if cross_db.check_table_exist(table_name) == False:
        cross_db.create_table(table_name, *columns)
    else:
        cross_db.delete_table(table_name)
        cross_db.create_table(table_name, *columns)
    
    for info in Data:
        cross_db._insert_data_single_dev(table_name, info)
    
    cross_db.conn.commit()                
    


if __name__ == "__main__":
    
    #create_Molecule_DB()
    
    create_Spectra_DB()
    
    
    