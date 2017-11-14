"""

This test code's purpose is to connect the entire simulation to pandexo simulation and determine detection


Here is the code Flow:


need a global cross section loading...

for each atmosphere type:
    for low/high clouds:
        simulate base spectra
        for each biosignature molecule:
            1. generate spectra with seas
            2. run pandexo with seas output
            3. run detection with pandexo output
                i. determine detection in feature "windows"
                ii. determine general detection location

"""

import os
import sys
import numpy as np
import random

from openpyxl.styles import PatternFill
from openpyxl import load_workbook, Workbook

DIR = os.path.abspath(os.path.dirname(__file__))
sys.path.insert(0, os.path.join(DIR, '../..'))

import SEAS_Main.simulation.transmission_spectra_simulator as theory
import SEAS_Main.simulation.observed_spectra_simulator as observe
import SEAS_Main.simulation.spectra_analyzer as analyze
from SEAS_Main.atmosphere_effects.biosig_molecule import biosig_interpolate,load_NIST_spectra

import SEAS_Utils as utils
import SEAS_Utils.common_utils.data_plotter as plotter
import SEAS_Utils.common_utils.configurable as config
from SEAS_Utils.common_utils.DIRs import *
from SEAS_Utils.common_utils.data_loader import NIST_Smile_List,NIST_to_HITRAN
from SEAS_Utils.common_utils.timer import simple_timer
from SEAS_Utils.common_utils.data_saver import save_npy, save_txt, Excel_Saver
from SEAS_Utils.common_utils.data_processor import merge_list_special

import SEAS_Aux.atmosphere_processes.TP_profile_generator as TPgen
import SEAS_Aux.atmosphere_processes.mixing_ratio_generator as mix

import pickle
from scipy import stats
import pandexo.engine.justdoit as jdi
import pandexo.engine.justplotit as jpi
import warnings
warnings.filterwarnings('ignore')


yellowFill = PatternFill(start_color='FFFF00',
                   end_color='FFFF00',
                   fill_type='solid')
redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')




def simulation_prep(atmosphere,type="File"):
    
    
    if type == "File":
        Filename = "Test_Earth"
        Molecules = atmosphere["Molecules"]
        MRFile = atmosphere["Mixing_Ratio_File"]
        TPFile = atmosphere["TP_File"]
        
        return Filename, Molecules, TPFile, MRFile


    Type                = atmosphere["Type"]
    Surface_Temperature = atmosphere["Surface_Temperature"]
    Molecules           = atmosphere["Molecules"]
    Mixing_Raio         = atmosphere["Ratio"]
    
    try:
        Filler = atmosphere["Filler"]
    except:
        Filler = "N2"
    Filename = "_".join([Type[:3],"_".join(Molecules),str(hash(str(atmosphere.values()))%10**8)])
    
    if Type == "isothermal":
        TP_input = config.Configuration("../../bin_stable/TP_Profile/TP_selection.cfg")["Test Isothermal Atmosphere"]
        TP_Name = "TP_%s.txt"%Filename
        if not os.path.isfile(os.path.join(TP_Profile_Data,TP_Name)):
            TP_simulator = TPgen.temperature_pressure_profile_generator(TP_input, name=TP_Name)
            TP_simulator.generate()
            TP_simulator.save()
    else:
        print "Non isothermal bulk simulation not implemented yet"
        sys.exit()
    
    # generate Mixing Ratio Files
    ratio_input = {}
    for i,m in enumerate(Molecules):
        ratio_input[m] = {}
        ratio_input[m]["Surface_Ratio"]  = Mixing_Raio[i]
        ratio_input[m]["End_Ratio"]      = None
        ratio_input[m]["Type"]           = "constant"
        ratio_input[m]["Transition"]     = None
        ratio_input[m]["Start_Pressure"] = None
        ratio_input[m]["End_Pressure"]   = None 
    
    MR_Name = "MR_%s.txt"%Filename
    if not os.path.isfile(os.path.join(Mixing_Ratio_Data,MR_Name)):
        simulator = mix.mixing_ratio_generator(ratio_input,
                                               filler=True,
                                               filler_molecule=Filler,
                                               pressures = [100000.0, 36800.0, 13500.0, 4980.0, 1830.0, 674.0, 248.0, 91.2, 33.5, 12.3, 4.54, 1.67, 0.614, 0.226, 0.0832, 0.0306, 0.0113, 0.00414, 0.00152, 0.00056, 0.000206, 7.58e-05, 2.79e-05, 1.03e-05],
                                               name=MR_Name,
                                               overwrite=True)
        simulator.generate()
        simulator.save()

    return Filename, Molecules,TP_Name, MR_Name

def set_base_parameters(user_input, Filename, TP_Name, MR_Name, 
                   cloud_type="grey", cloud_deck=10000, cloud_opacity=0.1):
    
    user_input["Simulation_Control"]["DB_DIR"]              = "Simulation_Band"
    user_input["Simulation_Control"]["DB_Name"]             = None#"cross_sec_Example.db"
    user_input["Simulation_Control"]["TP_Profile_Name"]     = TP_Name
    user_input["Simulation_Control"]["Mixing_Ratio_Name"]   = MR_Name

    user_input["Save"]["Intermediate_Data"]["cross_section_savename"] = "%s_Cross_Section.npy"%Filename
    
    user_input["Atmosphere_Effects"]["Cloud"]["model"]   = cloud_type
    user_input["Atmosphere_Effects"]["Cloud"]["deck"]    = cloud_deck
    user_input["Atmosphere_Effects"]["Cloud"]["opacity"] = cloud_opacity
    
    s   = theory.TS_Simulator(user_input)
    o   = observe.OS_Simulator(user_input)
    a   = analyze.Spectra_Analyzer(user_input)

    s.user_input = user_input
    
    return s,o,a

def simulate_theoratical_base_spectra(s,o,a):


    #create a base flat spectra based on planetary parameters
    s.Surface_g, s.Base_TS_Value = s.load_astrophysical_properties()
    
    # normalized pressure directly correspond to atmosphere layers.
    s.normalized_pressure = s.load_atmosphere_pressure_layers()
    
    # load mixing ration files and determine what molecules are added to the simulation
    # acquire the mixing ratio at each pressure (need to be interpolated to normalized pressure)
    s.normalized_molecules, s.MR_Pressure, s.molecule_MR = s.load_mixing_ratio()
    s.normalized_abundance = s.interpolate_mixing_ratio()
    
    # load temperature pressure profile
    s.TP_Pressure, s.TP_Temperature = s.load_TP_profile()        
    s.normalized_temperature = s.interpolate_TP_profile()
    
    # calculate the scale height for each layer of the atmosphere
    s.normalized_scale_height = s.calculate_scale_height()
    
    # load molecular cross section for main constituent of the atmosphere
    # will check first to see if the molecule is in the database 
    s.cross_db = s.check_molecules()
    s.nu, s.normalized_cross_section = s.load_molecule_cross_section()

    # load rayleigh scattering
    s.Rayleigh_enable = utils.to_bool(s.user_input["Atmosphere_Effects"]["Rayleigh"]["enable"])
    if s.Rayleigh_enable:
        s.normalized_rayleigh = s.load_rayleigh_scattering()   
    
    # load overlay
    s.Overlay_enable = utils.to_bool(s.user_input["Atmosphere_Effects"]["Overlay"]["enable"])
    if s.Overlay_enable:
        o_nu, o_xsec = s.load_overlay_effects()
        s.normalized_overlay = s.interpolate_overlay_effects(o_nu,o_xsec)

    # Reference spectra (w/o bio)
    #s.Reference_Transit_Signal = s.load_atmosphere_geometry_model()
    s.Base_Transit_Signal = s.load_atmosphere_geometry_model(Cloud=True)   

    s.nu_,s.ref_trans = o.calculate_convolve(s.nu, s.Base_Transit_Signal)
    
    # analyze the spectra
    s.wav_window = a.spectra_window(s.nu_,s.ref_trans,"T",0.2, 100.,s.deck_signal,"wav")

    print "Finish calculating Base Spectra"
    
    
    return s,o,a

def simulate_theoratical_biosignature_spectra(s,o,a, bio_molecule, source, bio_abundance):

    s.user_input["Atmosphere_Effects"]["Bio_Molecule"]["enable"] = True
    s.user_input["Atmosphere_Effects"]["Bio_Molecule"]["data_type"] = source
    s.user_input["Atmosphere_Effects"]["Bio_Molecule"]["molecule"] = bio_molecule
    s.user_input["Atmosphere_Effects"]["Bio_Molecule"]["is_smile"] = True
    
    # load biosignature molecules
    data_type     = s.user_input["Atmosphere_Effects"]["Bio_Molecule"]["data_type"]
    bio_molecule  = s.user_input["Atmosphere_Effects"]["Bio_Molecule"]["molecule"]
    s.is_smile    = utils.to_bool(s.user_input["Atmosphere_Effects"]["Bio_Molecule"]["is_smile"])
    s.nu, s.bio_cross_section = s.load_bio_molecule_cross_section(bio_molecule, data_type)
    s.bio_normalized_cross_section = np.concatenate([s.normalized_cross_section, s.bio_cross_section], axis=0)

    # for future. if filler, subtract from filler, else... do this
    s.bio_normalized_abundance = []
    for i,abundance in enumerate(s.normalized_abundance):
        s.bio_normalized_abundance.append([])
        for j in abundance:
            s.bio_normalized_abundance[i].append(j*(1-bio_abundance))
        s.bio_normalized_abundance[i].append(bio_abundance)
    s.bio_normalized_molecules = np.concatenate([s.normalized_molecules,[bio_molecule]], axis=0)

    s.Bio_Transit_Signal = s.load_atmosphere_geometry_model(bio=True)
    #s.Bio_Transit_Signal = s.load_atmosphere_geometry_model_with_cloud(cloud_deck,cloud_amount,bio=True)
    nu,bio_trans = o.calculate_convolve(s.nu, s.Bio_Transit_Signal)

    return nu,bio_trans




def run_pandexo_ETC(bin, tran, spectra, instrument):
    """
    manual changes are done to catalog.py and config.py in pandexo to make this code work
    jwst.py
    also change to locations.py in pysynphot
    """
        
    exo_dict = jdi.load_exo_dict()
    exo_dict['observation']['sat_unit'] = "%"
    exo_dict['observation']['sat_level'] = 80    #saturation level in percent of full well 
    exo_dict['observation']['noccultations'] = 10 #number of transits 
    exo_dict['observation']['R'] = None          #fixed binning. I usually suggest ZERO binning.. you can always bin later 
                                                 #without having to redo the calcualtion
    exo_dict['observation']['fraction'] = 1.0    #fraction of time in transit versus out = in/out
    exo_dict['observation']['noise_floor'] = 0   #this can be a fixed level or it can be a filepath 
    
    exo_dict['star']['type'] = 'phoenix'        #phoenix or user (if you have your own)
    exo_dict['star']['mag'] = 8.0               #magnitude of the system
    exo_dict['star']['ref_wave'] = 1.25         #For J mag = 1.25, H = 1.6, K =2.22.. etc (all in micron)
    exo_dict['star']['temp'] = 4000             #in K 
    exo_dict['star']['metal'] = 0.0             # as log Fe/H
    exo_dict['star']['logg'] = 4.0  

    #exo_dict['planet']['type'] = 'constant'
    #exo_dict['planet']['depth'] = 0.01                      #other options include "um","nm" ,"Angs", "secs" (for phase curves)
    exo_dict['planet']['exopath'] = spectra
    exo_dict['planet']['w_unit'] = "um"
    exo_dict['planet']['f_unit'] = "fp/f*"#'rp^2/r*^2'
    exo_dict['planet']['transit_duration'] = 2.0*60.0*60.0 
    
    if os.path.isfile("Temp_ETC_Sim.p"):
        out = pickle.load(open("Temp_ETC_Sim.p", 'rb'))
    else:
        out = jdi.run_pandexo(exo_dict, [instrument], output_file="Temp_ETC_Sim.p")
    
    if instrument == "MIRI LRS":
        ranges = [5,12]
    else:
        ranges = [1,5]
    

    m1,n1,x1,y1,e1 = jpi.jwst_1d_spec(out, R=bin, num_tran=tran, model=True, x_range=ranges)
    
    return m1,n1,x1,y1,e1

def determine_detection(s,o,a,molecule,bin_nu_MIRI, bin_ref_trans, bin_error):

    # setup detection window
    window = s.wav_window
    
    detected = [False]*len(window)
    detected_abundance = ["-"]*len(window)
    
    advance_detected = [False]*4
    one3sigma,three3sigma,one7sigma,three7sigma = "","","",""
    
    x1,y1 = load_NIST_spectra(molecule,["wl","T"],True)
    
    bio_data_min = min(x1)
    bio_data_max = max(x1)
    
    for ct,win in enumerate(window):
        low,high = win[0],win[1]
        if float(bio_data_min) > float(high) or float(bio_data_max) < float(low):
            detected_abundance[ct] = "/"
    
    for bio_abundance in np.array([0.001,0.002,0.005,0.01,0.02,0.05,
                                   0.1,0.2,0.5,1,2,5,10,20,50,100,
                                   1000,10000,100000])*10**-6:    
    
        nu,bin_bio_trans = simulate_theoratical_biosignature_spectra(s,o,a, molecule,"NIST",bio_abundance)            
        bin_means, bin_edges, binnumber = stats.binned_statistic(10000./nu[::-1], bin_bio_trans[::-1], bins=bin_nu_MIRI[0], range=(5,12))
        bin_width = (bin_edges[1] - bin_edges[0])
        bin_center = bin_edges[1:] - bin_width/2
    
        bin_center = np.array(bin_center)
        bin_ref_trans = np.array(bin_ref_trans)
        bin_bio_trans = np.array(bin_means)
        bin_error = np.array(bin_error)
        
        detection = bin_bio_trans - bin_ref_trans
        detect_error = bin_error*3
    
        detected_x = []
        detected_x_index = []
        detected_y1 = []
        detected_y2 = []
        for i,bin in enumerate(detection-detect_error):
            if bin > 0 and detection[i]>0:
                detected_x.append(bin_center[i])
                detected_x_index.append(i)
                detected_y1.append(bin_ref_trans[i])
                detected_y2.append(bin_bio_trans[i])
                
        # merge adjacent detection ( within 5 bins from each other?)
        merged_detected_x = merge_list_special(detected_x_index,bin_ref_trans,bin_bio_trans,bin_error)
        feature_x = np.array(map(bin_center.__getitem__, merged_detected_x))
        feature_1 = np.array(map(bin_ref_trans.__getitem__, merged_detected_x))
        feature_2 = np.array(map(bin_bio_trans.__getitem__, merged_detected_x))
        feature_error = np.array(map(bin_error.__getitem__,merged_detected_x))
        feature_detection = (feature_2-feature_1)/feature_error
    
        #print feature_x
        #print feature_detection
        
        
        
        """
        plt.title("Simulated Exoplanet Atmosphere Detection with %s bins and %s transits at 3 Sigma")
        plt.xlabel('Wavelength [microns]')
        plt.ylabel('fp/f*') 
    
        plt.errorbar(bin_center,bin_ref_trans,bin_error)
        plt.plot(bin_center,bin_bio_trans)   
        plt.plot(detected_x,detected_y1,".")
        plt.plot(detected_x,detected_y2,".")
        plt.show()
        
        """
    
        # 3 sigma detectable in defined window?

        for k,win in enumerate(window):
            low,high = float(win[0]),float(win[1])
            for x in feature_x:
                if detected[k] == False:
                    if x> low and x< high:
                        detected_abundance[k] = float("%.2g"%(bio_abundance*10**6))
                        detected[k] = True
                        
            

        #print detected
        #print detected_abundance

    # minimum abundance needed for one 3 sigma detection
    
        detected_number = len(feature_detection)
        
        if detected_number>0 and advance_detected[0]==False:
            one3sigma = float("%.2g"%(bio_abundance*10**6))
            advance_detected[0]= True
    
    # minimum abundance needed for three 3 sigma detection
    
        if detected_number>3 and advance_detected[1]==False:
            three3sigma = float("%.2g"%(bio_abundance*10**6))    
            advance_detected[1]= True
    
    # minimum abundance needed for one 7 sigma detection
    
        if sum(np.array(feature_detection)>7) >0 and advance_detected[2]==False:
            one7sigma = float("%.2g"%(bio_abundance*10**6))    
            advance_detected[2]= True
    
    # minimum abundance needed for one 7 sigma detection
    
        if sum(np.array(feature_detection)>7) >0 and advance_detected[3]==False:
            three7sigma = float("%.2g"%(bio_abundance*10**6))    
            advance_detected[3]= True

        #print one3sigma,three3sigma,one7sigma,three7sigma


    detection_information  = [one3sigma,three3sigma,one7sigma,three7sigma],detected_abundance,advance_detected[0]
    
    return detection_information



def Third_Stage_Simulation():

    Timer = simple_timer(4)

    user_input = config.Configuration("../../bin_stable/a.Main/user_input_dev.cfg")
    atmo_input = config.Configuration("../../bin_stable/a.Main/atmosphere_prototype.cfg")
    atmosphere_types = atmo_input["Realistic"]

    molecule_type = "NIST"
    info = NIST_Smile_List()
    molecule_smiles = info[0]
    formula = info[3]
    IUPAC_name = info[4]
    
    
    total_simulation = len(atmosphere_types)*len(molecule_smiles)

    # create a workbook and fill in the general page
    name = "Third_Stage_NIST_Detectable_Abundance_Earth_Grey_Cloud_JWST_MIRI.xlsx"
    path = os.path.join(Result_DIR,"Third_Stage_Simulation")
    sheet = "General Result"
    
    WB = Excel_Saver(path, name)
    
    WB.load_sheet(sheet)
    try:
        WB.delete_sheet("Sheet")
    except:
        pass
    WB.write_column_header(molecule_smiles)
    WB.write_column(formula,column_offset=1,row_offset=2)    
    WB.write_column(IUPAC_name,column_offset=2,row_offset=2)    
    WB.input_data["A1"].value = "Smiles"
    WB.input_data["B1"].value = "Formula"
    WB.input_data["C1"].value = "IUPAC"
    WB.write_row_header(atmosphere_types.keys(),offset=2)


    
    for i, atmosphere in enumerate(atmosphere_types):

        print "Simulating '%s' Atmosphere"%atmosphere
        
        sheetname = atmosphere
        atmosphere_input_data = WB.load_sheet(sheetname,False)
        
        
        Filename, Molecules, TP_Name, MR_Name = simulation_prep(atmosphere_types[atmosphere])
    
        s,o,a = set_base_parameters(user_input, Filename, TP_Name, MR_Name, 
                   cloud_type="grey", cloud_deck=10000, cloud_opacity=0.1)
        
        # reference spectra
        s,o,a = simulate_theoratical_base_spectra(s,o,a)
        
        
        save_txt("","temp_ref_trans.txt",np.array([10000./s.nu_[::-1],s.ref_trans[::-1]]).T)

        
        bins = 100
        tran = 100
        
        model_x, model_y, bin_nu_MIRI, bin_ref_trans_MIRI, bin_ref_error_MIRI = run_pandexo_ETC(bin, tran, "temp_ref_trans.txt", "MIRI LRS")
        
        bin_ref = bin_ref_trans_MIRI[0][:-1]
        error = bin_ref_error_MIRI[0][:-1]
        """
        bin_nu_NIRS, bin_ref_trans_NIRS, bin_ref_error_NIRS = run_pandexo_ETC(s.nu_,s.ref_trans, bin, tran, "NIRSpec")
        
        bin_nu = bin_nu_NIRS + bin_nu_MIRI 
        bin_ref_trans = bin_ref_trans_NIRS+bin_ref_trans_MIRI
        bin_error = bin_ref_error_NIRS + bin_ref_error_MIRI
        """
        
        atmosphere_input_data["A1"].value = "Smiles"
        atmosphere_input_data["B1"].value = "Formula" 
        atmosphere_input_data["C1"].value = "IUPAC" 
        atmosphere_input_data["D1"].value = "1 3sigma" 
        atmosphere_input_data["E1"].value = "3 3sigma" 
        atmosphere_input_data["F1"].value = "1 7sigma" 
        atmosphere_input_data["G1"].value = "3 7sigma" 
        
        for k,win in enumerate(s.wav_window):
            col = chr(ord("H")+k)
            atmosphere_input_data["%s%d"%(col, 1)].value = "-".join(["%.2f"%(float(win[0])),"%.2f"%(float(win[1]))])



        for j, molecule in enumerate(molecule_smiles):
        
            """
            if j >= 10:
                break
            """
            
            if j%5 ==0:
                
                passed = i*len(molecule_smiles)+j
                
                print Timer.progress(passed,total_simulation)
            
            
            
            # need NIST smiles to HITRAN Formulas
            if NIST_to_HITRAN(formula[j]) in Molecules:
                detection_information = ["IN"]
            else:
                detection_information = determine_detection(s,o,a, molecule, bin_nu_MIRI, bin_ref, error)

            #print molecule, detection_information
    
    
            atmosphere_input_data["%s%d"%("A", j+2)].value = molecule
            atmosphere_input_data["%s%d"%("B", j+2)].value = formula[j]  
            atmosphere_input_data["%s%d"%("C", j+2)].value = IUPAC_name[j]  
    
            if detection_information != ["IN"]:
                
                
                for a,abundance in enumerate(detection_information[0]):
                    det_col = chr(ord("D")+i+a)
                    atmosphere_input_data["%s%d"%(det_col, j+2)].value = abundance
                    if abundance < 100:
                        atmosphere_input_data["%s%d"%(det_col, j+2)].fill = yellowFill
                    if abundance < 1:
                        atmosphere_input_data["%s%d"%(det_col, j+2)].fill = redFill   
                               
    
                
                for b,abundance in enumerate(detection_information[1]):
                    win_col = chr(ord("H")+i+b)
                    atmosphere_input_data["%s%d"%(win_col, j+2)].value = abundance
                    if abundance < 100:
                        atmosphere_input_data["%s%d"%(win_col, j+2)].fill = yellowFill
                    if abundance < 1:
                        atmosphere_input_data["%s%d"%(win_col, j+2)].fill = redFill     
    
                detected = detection_information[2]
                WB.input_data["%s%d"%("D", j+2)].value = detected
                if detected:
                    WB.input_data["%s%d"%("D", j+2)].fill = yellowFill    
            else:
                WB.input_data["%s%d"%("D", j+2)].value = "In"
                
                
            # bin to jwst bins.
            # find out jwst bin method.
            #bin_nu, bin_bio_trans = nu, bio_trans
    
            #bin_nu_MIRI, bin_bio_trans_MIRI = run_pandexo_ETC(nu,bio_trans, "MIRI LRS")
            #bin_nu_NIRS, bin_bio_trans_NIRS = run_pandexo_ETC(nu,bio_trans, "NIRSpec")    
            
            #bin_bio_trans = bin_bio_trans_NIRS+bin_bio_trans_MIRI
            
            #detection_information = determine_detection(bin_nu, bin_ref_trans, bin_bio_trans, bin_error)
            
            WB.save()
    WB.save()
    print Timer.total()
    return "Done" 
    

if __name__ == "__main__":
    
    
    Third_Stage_Simulation()
    







