"""
code to sort all molecules

"""


import os
import sys
import numpy as np
import random

import matplotlib.pyplot as plt

from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl import load_workbook, Workbook

DIR = os.path.abspath(os.path.dirname(__file__))
sys.path.insert(0, os.path.join(DIR, '../..'))


from SEAS_Utils.common_utils.colors import *
import SEAS_Utils.common_utils as utils
import SEAS_Utils.common_utils.configurable as config
from SEAS_Utils.common_utils.DIRs import *
from SEAS_Utils.common_utils.data_loader import NIST_Smile_List,ASM_Smile_List,HITRAN_to_NIST

from rdkit import Chem
from rdkit.Chem import rdFMCS

def molecule_bond(smiles,bond):
    
    m1 = Chem.MolFromSmiles(smiles)
    m2 = Chem.MolFromSmiles(bond)
    
    if m1==m2:
        return False
    return m1.HasSubstructMatch(m2)



def bonds_in_molecule():
    
    datapath = os.path.join(molecule_info,"bond","all_bond.txt")
    all_bond = open(datapath,"r").read().split("\n")
    
    molecular_info = NIST_Smile_List()
    
    Smiles = molecular_info[0]
    Formula = molecular_info[3]
    IUPAC = molecular_info[4]
    

    for smiles in Smiles:
        print smiles,
        for bond in all_bond:

            if molecule_bond(smiles,bond):
                print bond,
        print



def bond_not_detected():


    name = "Fourth_Stage_NIST_Detectable_Abundance_Proposed_Grey_Cloud_JWST_dfix.xlsx"
    path = os.path.join(Result_DIR,"Fourth_Stage_Simulation")    
    
    WBI = load_workbook(os.path.join(path,name))
    
    atmo_input = config.Configuration("../../bin_stable/a.Main/atmosphere_prototype.cfg")
    atmosphere_types = atmo_input["Proposed"]




    bond_path = os.path.join(molecule_info,"bond","all_bond.txt")
    all_bond = open(bond_path,"r").read().split("\n")


    molecular_info = NIST_Smile_List()
    Smiles = molecular_info[0]

    for atmosphere in atmosphere_types:
        
        Molecules = " ".join(atmosphere_types[atmosphere]["Molecules"])
        Abundance = " ".join(atmosphere_types[atmosphere]["Ratio"])  
        
        detectability_sheet = WBI[atmosphere]
        
        detectSmiles = []
        num = 2+434
        for i in range(100):
            
            smile = detectability_sheet["A%d"%num].value
            
            if smile == None:
                break
            
            detectSmiles.append(smile)
            num+=1
            
            
    
        #detectpath = "temp_faor.txt"
        #detectSmiles = open(detectpath,"r").read().split("\n")
        
    
        
        def get_data(Input_Smiles, sort = True, input=""):
            
            frequency = {}
            for bond in all_bond:
                frequency[bond] = 0
                 
            for smiles in Input_Smiles:
                for bond in all_bond:
                    if molecule_bond(smiles,bond):
                        frequency[bond] +=1
            
            bonds = []
            occurance = []
            
            if sort:
                for key, value in sorted(frequency.iteritems(), key=lambda (k,v): (v,k), reverse=True):
                    if value != 0:
                        print key,value
                        bonds.append(key)
                        occurance.append(value)
            else:
                for key in input:
                    value = frequency[key]
                    if value !=0:
                        print key,value
                        bonds.append(key)
                        occurance.append(value)            
            
            x = np.arange(len(occurance))  
            
            return x,occurance,bonds
        
        x,occurance,bonds = get_data(Smiles)
    
        x1,occurance1,bonds1 = get_data(detectSmiles,False,bonds)
        
        plt.figure(figsize=(16,8))
        plt.title("Number of molecules that has the following molecular bond (NIST vs Bottom 100) in %s\n%s\n%s"%(atmosphere,Molecules,Abundance))
        plt.xlabel("Bonds that Appear\n Black: NIST. White: Not Detected. Yellow: Occurance Factor (Not Detected/100)/(NIST/534)")
        plt.ylabel("Occurance Rate")
        
        rects1 = plt.bar(x,occurance,log=True,color="b")
        rects2 = plt.bar(x1,occurance1,log=True,color="r")
        
   
        
        
        for ddd,rect in enumerate(rects2):
            height = rect.get_height()
            
            plt.text(rect.get_x() + rect.get_width()/2., 1.05*height,
                    '%d' % int(height),ha='center', va='bottom', color="w")
            
            Rect = rects1[ddd]
            Total_height = Rect.get_height()
 
 
            plt.text(Rect.get_x() + Rect.get_width()/2., 1.05*Total_height,
                    '%d' % int(Total_height),ha='center', va='bottom') 
            
            occurance_factor = "%.3g"%(5.34*height/float(Total_height))
            
            plt.text(Rect.get_x() + Rect.get_width()/2., 0.8,
                    occurance_factor,ha='center', va='bottom',color="y")
        
        
        
        plt.legend((rects1[0],rects2[0]),("NIST","Not Detected"))
        
        plt.xticks(x,bonds)
        
        plt.savefig(os.path.join(path,"not_molecule_bond_compare_NIST_%s.png"%atmosphere))
        plt.show()


def bond_detected():


    name = "Fourth_Stage_NIST_Detectable_Abundance_Proposed_Grey_Cloud_JWST_dfix.xlsx"
    path = os.path.join(Result_DIR,"Fourth_Stage_Simulation")    
    
    WBI = load_workbook(os.path.join(path,name))
    
    atmo_input = config.Configuration("../../bin_stable/a.Main/atmosphere_prototype.cfg")
    atmosphere_types = atmo_input["Proposed"]




    bond_path = os.path.join(molecule_info,"bond","all_bond.txt")
    all_bond = open(bond_path,"r").read().split("\n")


    molecular_info = NIST_Smile_List()
    Smiles = molecular_info[0]

    for atmosphere in atmosphere_types:
        
        Molecules = " ".join(atmosphere_types[atmosphere]["Molecules"])
        Abundance = " ".join(atmosphere_types[atmosphere]["Ratio"])  
        
        detectability_sheet = WBI[atmosphere]
        
        detectSmiles = []
        num = 2
        for i in range(100):
            detectSmiles.append(detectability_sheet["A%d"%num].value)
            num+=1
            
            
    
        #detectpath = "temp_faor.txt"
        #detectSmiles = open(detectpath,"r").read().split("\n")
        
    
        
        def get_data(Input_Smiles, sort = True, input=""):
            
            frequency = {}
            for bond in all_bond:
                frequency[bond] = 0
                 
            for smiles in Input_Smiles:
                for bond in all_bond:
                    if molecule_bond(smiles,bond):
                        frequency[bond] +=1
            
            bonds = []
            occurance = []
            
            if sort:
                for key, value in sorted(frequency.iteritems(), key=lambda (k,v): (v,k), reverse=True):
                    if value >= 5:
                        print key,value
                        bonds.append(key)
                        occurance.append(value)
            else:
                for key in input:
                    value = frequency[key]
                    if value >= 5:
                        print key,value
                        bonds.append(key)
                        occurance.append(value)            
            
            x = np.arange(len(occurance))  
            
            return x,occurance,bonds
        
        x,occurance,bonds = get_data(Smiles)
    
        x1,occurance1,bonds1 = get_data(detectSmiles,False,bonds)
        
        plt.figure(figsize=(16,8))
        plt.title("Number of molecules that has the following molecular bond (NIST vs Top 100) in %s\n%s\n%s"%(atmosphere,Molecules,Abundance))
        plt.xlabel("Bonds that Appear\n Black: NIST. White: Detected. Yellow: Occurance Factor (Detected/100)/(NIST/534)")
        plt.ylabel("Occurance Rate")
        
        rects1 = plt.bar(x,occurance,log=True,color="b")
        rects2 = plt.bar(x1,occurance1,log=True,color="g")
        
   
        
        
        for ddd,rect in enumerate(rects2):
            height = rect.get_height()
            
            plt.text(rect.get_x() + rect.get_width()/2., 1.05*height,
                    '%d' % int(height),ha='center', va='bottom', color="w")
            
            Rect = rects1[ddd]
            Total_height = Rect.get_height()
 
 
            plt.text(Rect.get_x() + Rect.get_width()/2., 1.05*Total_height,
                    '%d' % int(Total_height),ha='center', va='bottom') 
            
            occurance_factor = "%.3g"%(5.34*height/float(Total_height))
            
            plt.text(Rect.get_x() + Rect.get_width()/2., 0.8,
                    occurance_factor,ha='center', va='bottom',color="y")
        
        
        
        plt.legend((rects1[0],rects2[0]),("NIST","Detected"))
        
        plt.xticks(x,bonds)
        
        plt.savefig(os.path.join(path,"molecule_bond_compare_NIST_%s.png"%atmosphere))
        plt.show()
    


def bond_compare_top_bottom_100():
    

    name = "Fourth_Stage_NIST_Detectable_Abundance_Proposed_Grey_Cloud_JWST_dfix.xlsx"
    path = os.path.join(Result_DIR,"Fourth_Stage_Simulation")    
    
    WBI = load_workbook(os.path.join(path,name))
    
    atmo_input = config.Configuration("../../bin_stable/a.Main/atmosphere_prototype.cfg")
    atmosphere_types = atmo_input["Proposed"]




    bond_path = os.path.join(molecule_info,"bond","all_bond.txt")
    all_bond = open(bond_path,"r").read().split("\n")


    molecular_info = NIST_Smile_List()
    Smiles = molecular_info[0]

    for atmosphere in atmosphere_types:
        
        Molecules = " ".join(atmosphere_types[atmosphere]["Molecules"])
        Abundance = " ".join(atmosphere_types[atmosphere]["Ratio"])  
        
        detectability_sheet = WBI[atmosphere]
        
        detectSmiles = []
        notdetectSmiles = []
        
        
        num = 2
        for i in range(100):
            detectSmiles.append(detectability_sheet["A%d"%num].value)
            num+=1
        
        num = 2+434
        for i in range(100):
            
            smile = detectability_sheet["A%d"%num].value
            
            if smile == None:
                break
            
            notdetectSmiles.append(smile)
            num+=1
            
        
    
        #detectpath = "temp_faor.txt"
        #detectSmiles = open(detectpath,"r").read().split("\n")
        
    
        
        def get_data(Input_Smiles, sort = True, input=""):
            
            frequency = {}
            for bond in all_bond:
                frequency[bond] = 0
                 
            for smiles in Input_Smiles:
                for bond in all_bond:
                    if molecule_bond(smiles,bond):
                        frequency[bond] +=1
            
            bonds = []
            occurance = []
            
            if sort:
                for key, value in sorted(frequency.iteritems(), key=lambda (k,v): (v,k), reverse=True):
                    if value >= 2:
                        print key,value
                        bonds.append(key)
                        occurance.append(value)
            else:
                for key in input:
                    value = frequency[key]
                    if value != 0:
                        print key,value
                        bonds.append(key)
                        occurance.append(value)            
            
            x = np.arange(len(occurance))  
            
            return x,occurance,bonds
        
        x,occurance,bonds = get_data(Smiles)
    
        x1,occurance1,bonds1 = get_data(detectSmiles,False,bonds)
        
        x2,occurance2,bonds2 = get_data(notdetectSmiles,False,bonds)
        
        plt.figure(figsize=(16,8))
        plt.title("Number of molecules that has the following molecular bond (NIST vs Top/Bottom 100) in %s\n%s\n%s"%(atmosphere,Molecules,Abundance))
        plt.xlabel("Bonds that Appear\n Black: NIST. White: Detected. Yellow: Not Detected.")
        plt.ylabel("Occurance Rate")
        
        
        bar_width = 0.4
        rects1 = plt.bar(x,occurance,bar_width*2,log=True,color="b")
        rects2 = plt.bar(x1-bar_width/2,occurance1,bar_width,log=True,color="g")
        rects3 = plt.bar(x2+bar_width/2,occurance2,bar_width,log=True,color="r")
   
        
        
        for ddd,rect in enumerate(rects1):
            Total_height = rect.get_height()
            plt.text(rect.get_x() + rect.get_width()/2., 1.05*Total_height,
                    '%d' % int(Total_height),ha='center', va='bottom') 
            
            try:
                Rect = rects2[ddd]
                height = Rect.get_height()
                plt.text(Rect.get_x() + Rect.get_width()/2., 1.05*height,
                        '%d' % int(height),ha='center', va='bottom', color="w")
            except:
                pass
                
            
            try:
                NotRect = rects3[ddd]
                Not_height = NotRect.get_height()
                plt.text(NotRect.get_x() + NotRect.get_width()/2., 1.05*Not_height,
                        '%d' % int(Not_height),ha='center', va='bottom', color="y")            
            except:
                pass

            
            #occurance_factor = "%.3g"%(5.34*height/float(Total_height))
            
            #plt.text(Rect.get_x() + Rect.get_width()/2., 0.8,
            #        occurance_factor,ha='center', va='bottom',color="y")
            
        
        plt.legend((rects1[0],rects2[0],rects3[0]),("NIST","Detected","Not Detected"))
        
        plt.xticks(x,bonds)
        
        plt.savefig(os.path.join(path,"compare_molecule_bond_compare_NIST_%s.png"%atmosphere))
        plt.show()    


    

if __name__ == "__main__":
    
    #main()
    
    
    bond_compare_top_bottom_100()











